import openpyxl

path="C:/Users/Admin/Desktop/Learning/data driven testing/write.xlsx"

workbook=openpyxl.load_workbook(path)
#When there are more than 1 sheets
sheet=workbook["Sheet1"]

#when only 1 sheet is there
#sheet=workbook.active

#---------------------------Same data------------------------#
rows = sheet.max_row

columns = sheet.max_column
p=sheet.cell(row=3,column=1)
for r in range(1,6):
    for c in range(1,4):
        sheet.cell(row=1, column=4).value = "Email"
        #save file
workbook.save(path)


