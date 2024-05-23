#Read file from excel

import openpyxl

#Get excel file path with /name and also we need to mention sheet name
path="C:/Users/Admin/Desktop/Book1.xlsx"

workbook=openpyxl.load_workbook(path)

# to identify active worksheet
Active = workbook.active
# to identify maximum rows count
row=(Active.max_row)
print(row)
# to identify maximum columns count
col=(Active.max_column)
print(col)
#printing of paritucular value from cell
#c=Active.cell(row=3,column=1)
# to retrieve the cell value and print
#print(c.value)


#to print all values from rows and coloums
#using for loop
# iterate till the count of occupied rows
for m in range(1,Active.max_row + 1):
# iterate till the count of occupied columns
    for n in range(1,Active.max_column + 1):
#to get the cell data and print
      #print(Active.cell(row=m,column=n).value)
      #print()  #to print in separate line
      print(Active.cell(row=m,column=n).value,end='                       ') #To print in table format