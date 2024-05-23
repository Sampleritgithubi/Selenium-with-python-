#All excel utils
#this can be used in data driven testing
import openpyxl
from openpyxl.styles import PatternFill

#check no of rows
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)


